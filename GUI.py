import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from pathlib import Path
import firebase_admin
from firebase_admin import credentials
from firebase_admin import storage
import os
from zipfile import ZipFile
import cv2
from cursorClass import *
from imutils import resize
from PIL import Image
from PIL import ImageTk
import feature_extraction
from feature_extraction import Parkinson_movements
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg)
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("ParkinsonApp")
        self.geometry("1340x768")  # setting the window's size
        self.configure(bg="#3A7FF6")

        self.tabControl = ttk.Notebook(self)


class Tab1(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)
        # Create app canvas (rectangles, titles, layout)
        self.canvas = tk.Canvas(self, bg="#3A7FF6", height=720, width=1280, bd=0, highlightthickness=0, relief="ridge")
        self.canvas.place(x=0, y=0)

        # Variables
        self.folder_var = tk.StringVar(value="Carpeta no seleccionada")
        self.patientID_var = tk.StringVar(value='')
        self.movement_var = tk.StringVar(value="Seleccione movimiento")
        self.hand_var = tk.StringVar(value="Seleccione mano")
        self.test_var = tk.StringVar(value="Seleccione prueba")
        self.updrs_var = tk.IntVar(value='UPDRS')
        self.process_var = tk.StringVar(value="")

        self.video_cap = None
        self.video_fps = 30
        self.video_ts = 1 // 30
        self.video_pause = False
        self.mov_eval = ''
        self.csv_path = ''

        # Assets
        self.entry_img = tk.PhotoImage(file=relative_to_assets("entry.png"))
        self.btn_search_img = tk.PhotoImage(file=relative_to_assets("btn_search.png"))
        self.bg_dropdown_img = tk.PhotoImage(file=relative_to_assets("bg_dropdown.png"))
        self.btn_process_img = tk.PhotoImage(file=relative_to_assets("btn_process.png"))
        self.btn_updrs_img = tk.PhotoImage(file=relative_to_assets("btn_updrs.png"))
        self.btn_vidplay_img = tk.PhotoImage(file=relative_to_assets("btn_play.png"))
        self.btn_vidpause_img = tk.PhotoImage(file=relative_to_assets("btn_pause.png"))
        self.btn_vidfirst_img = tk.PhotoImage(file=relative_to_assets("btn_first.png"))

        self.create_widgets()

    def create_widgets(self):
        # Title
        self.canvas.create_text(37.0, 37.0, anchor="nw", text="Seleccionar prueba", fill="#FFFFFF",
                                font=("Roboto Bold", 36 * -1))

        # Carpeta de descarga: text, entry, button
        self.canvas.create_text(55.0, 80.0, anchor="nw", text="Carpeta de descarga", fill="#FFFFFF",
                                font=("Roboto", 24 * -1))
        self.canvas.create_image(210.0, 140, image=self.entry_img)
        folder_label = tk.Label(self, textvariable=self.folder_var, bd=0, bg="#FFFFFF", fg="#000000",
                                highlightthickness=0, anchor="w", font=("Roboto", 13 * -1))
        folder_label.place(x=62.0, y=128.5, width=290.0)
        folder_btn = tk.Button(self, image=self.btn_search_img,
                               borderwidth=0,
                               highlightthickness=0,
                               command=self.get_selected_folder,
                               relief="flat")
        folder_btn.place(x=383.0, y=120.0, width=137.0, height=47.0)

        # ID del paciente: text, entry, button
        self.canvas.create_text(55.0, 185.0, anchor="nw", text="ID del paciente", fill="#FFFFFF",
                                font=("Roboto", 24 * -1))
        self.canvas.create_image(210, 245, image=self.entry_img)
        patient_id = tk.Entry(self, textvariable=self.patientID_var, bd=0, bg="#FFFFFF", fg="#000000",
                              highlightthickness=0)
        patient_id.place(x=62.0, y=222.5, width=275.0, height=45.0)
        search_btn = tk.Button(self, image=self.btn_search_img,
                               borderwidth=0,
                               highlightthickness=0,
                               command=self.get_tests,
                               relief="flat")
        search_btn.place(x=383.0, y=225.0, width=137.0, height=47.0)

        # Seleccionar: rectangle background, 3 dropdowns, process video button, UPDRS classification (btn and dropdown)
        self.canvas.create_text(55.0, 320.0, anchor="nw", text="Seleccionar video", fill="#FFFFFF",
                                font=("Roboto Bold", 24 * -1))
        self.canvas.create_image(286.0, 480.0, image=self.bg_dropdown_img)

        hand_entry = tk.OptionMenu(self, self.hand_var, "Derecha", "Izquierda")
        hand_entry.place(x=70, y=380.0, width=250.0, height=30.0)
        hand_entry.config(bg='#FFFFFF', fg='#000000', font=("Roboto Bold", 16 * -1))

        movement_entry = tk.OptionMenu(self, self.movement_var, "Golpeteo de dedos", "Prono supinacion",
                                       "Cierre de puño")
        movement_entry.place(x=70, y=440.0, width=250.0, height=30.0)
        movement_entry.config(bg='#FFFFFF', fg='#000000', font=("Roboto Bold", 16 * -1))

        self.test_entry = tk.OptionMenu(self, self.test_var, 'No hay pruebas disponibles')
        self.test_entry.place(x=70, y=500, width=250.0, height=30.0)
        self.test_entry.config(bg='#FFFFFF', fg='#000000', font=("Roboto Bold", 16 * -1))

        updrs_entry = tk.OptionMenu(self, self.updrs_var, '0', '1', '2', '3', '4')
        updrs_entry.place(x=380, y=440, width=80.0, height=30.0)
        updrs_entry.config(bg='#FFFFFF', fg='#000000', font=("Roboto Bold", 16 * -1))

        process_btn = tk.Button(self, image=self.btn_process_img, borderwidth=0, highlightthickness=0,
                                command=self.download_test, relief="flat")
        process_btn.place(x=80.0, y=620.0, width=200.0, height=47.0)

        self.updrs_btn = tk.Button(self, text="Clasicación UPDRS", borderwidth=0, highlightthickness=0,
                                   command=self.updrs_clasification, relief="flat", state=tk.DISABLED)
        self.updrs_btn.place(x=320.0, y=620.0, width=200.0, height=47.0)

        # Process Label
        self.process_label = tk.Label(self, textvariable=self.process_var, bd=0, bg="#FFFFFF", fg="#888888",
                                      highlightthickness=0, anchor="w", font=("Roboto", 16 * -1))
        self.process_label.place(x=80, y=550)

        # Gray rectangle where video will play
        self.canvas.create_rectangle(640.0, 0.0, 1280.0, 720.0, fill="#C4C4C4", outline="")
        self.video_label = tk.Label(self, bg="#C4C4C4")
        self.video_label.place(x=750.0, y=10)
        # Video control buttons
        self.playvid_btn = tk.Button(self, image=self.btn_vidplay_img, borderwidth=0, highlightthickness=0,
                                     command=self.vid_play, relief="flat")
        self.playvid_btn.place(x=787.0, y=640.0, width=50.0, height=50.0)
        pausevid_btn = tk.Button(self, image=self.btn_vidpause_img, borderwidth=0, highlightthickness=0,
                                 command=self.vid_pause, relief="flat")
        pausevid_btn.place(x=935.0, y=640.0, width=50.0, height=50.0)
        firstvid_btn = tk.Button(self, image=self.btn_vidfirst_img, borderwidth=0, highlightthickness=0,
                                 command=self.vid_first, relief="flat")
        firstvid_btn.place(x=1083.0, y=640.0, width=50.0, height=50.0)

    def get_selected_folder(self):
        # Prompt for user to choose folder for downloads
        selected_folder = filedialog.askdirectory()
        self.folder_var.set(selected_folder)

    def get_tests(self):
        # Get list of tests for selected patient ID stored in Firebase and shows them in test dropdoen
        global bucket
        blobs = bucket.list_blobs(prefix=self.patientID_var.get() + '/')
        # Reset dropdown menus
        self.hand_var.set("Seleccione mano")
        self.movement_var.set("Seleccione movimiento")
        self.test_var.set("Seleccione fecha")
        self.updrs_var.set("UPDRS")
        self.process_var.set("")

        # Get test names for patient ID
        test_list = []
        for blob in blobs:
            folder_name = blob.name
            start_idx = folder_name.find('/')
            test_name = folder_name[start_idx + 1:start_idx + 17]
            status = "ON" if test_name[-1] == "N" else "OFF"
            # Get test name formatted
            test_list.append(test_name[0:2] + "-" + test_name[2:4] + "-" + test_name[4:8] +
                             ", " + test_name[9:11] + ":" + test_name[11:13] + ", " + status)

        self.test_entry['menu'].delete(0, 'end')
        for choice in test_list:
            self.test_entry['menu'].add_command(label=choice, command=tk._setit(self.test_var, choice))

    def download_test(self):
        # Download test folder from Firebase
        # global bucket
        download_folder = self.folder_var.get()
        pID = self.patientID_var.get()
        test = self.test_var.get()

        # Check all fields have values
        if(download_folder == "Carpeta no seleccionada" or download_folder == ""):
            messagebox.showinfo(message="Debe seleccionar una carpeta de descarga.", title="Seleccionar Carpeta")
            return
        if ('Seleccione' in (test or self.hand_var.get() or self.movement_var.get())):
            messagebox.showinfo(message="Debe seleccionar una opción en todos los campos.", title="Seleccionar opciones")
            return

        self.process_var.set("Procesando...")

        # Download test files from firebase, but first check if folder already exists
        folder_path = pID + '/' + test
        folder_path = folder_path.replace(':', '-')

        if not os.path.isdir(download_folder + '/' + pID):
            os.mkdir(download_folder + '/' + pID)
        if not os.path.isdir(download_folder + '/' + folder_path):
            os.mkdir(download_folder + '/' + folder_path)
            firebase_path = pID + '/' + test[0:2] + test[3:5] + test[6:10] + "_" + test[12:14] + test[15:17] + "_" + test[19:21]
            blob = bucket.blob(firebase_path + '/test.zip')
            blob.download_to_filename(download_folder + '/' + folder_path + "/test.zip")
            # Extract all files to folder
            with ZipFile(download_folder + '/' + folder_path + "/test.zip", 'r') as f:
                #f.printdir()
                f.extractall(download_folder + '/' + folder_path)
            # Delete ZIP
            os.remove(download_folder + '/' + folder_path + "/test.zip")

        self.mov_eval, hand_name = get_moveval(self.movement_var.get(), self.hand_var.get())
        video_path = download_folder + '/' + folder_path + '/' + self.mov_eval + hand_name + '.mp4'
        self.csv_path = download_folder + '/' + folder_path + '/' + self.mov_eval + hand_name + '.csv'

        self.video_cap = cv2.VideoCapture(video_path)

        # Prueba
        #self.mov_eval = 'fingertap'
        #video_path = '/Users/santiagorojasjaramillo/Desktop/Prueba_1234/fingertap_r.mp4'
        # self.csv_path = '/Users/santiagorojasjaramillo/Desktop/mov_csv/fingertap_r.csv'
        #self.video_cap = cv2.VideoCapture(video_path)
        # Prueba

        self.video_fps = self.video_cap.get(cv2.CAP_PROP_FPS)
        self.video_ts = int((1 / self.video_fps) * 1000)

        self.vid_first_frame()
        self.updrs_btn['state'] = tk.NORMAL

        self.get_trajectory(video_path)

    def updrs_clasification(self):
        escala = self.updrs_var.get()
        xlsx_folder = self.folder_var.get()
        pID = self.patientID_var.get()
        test = self.test_var.get()
        hand = self.hand_var.get()
        filas_eliminadas = 2
        nombre_xlsx = xlsx_folder + "/ClasificacionVideos.xlsx"
        if not os.path.exists(nombre_xlsx):
            wb = Workbook()
            # grab the active worksheet
            ws = wb.active
            fila_max_analizada = ws.max_row
            ws['A1'] = 'ID del paciente'
            ws['B1'] = 'Movimiento evaluado'
            ws['C1'] = 'Mano'
            ws['D1'] = 'Fecha de la prueba'
            ws['E1'] = 'Clasificación UPDRS'
            ws.append([str(pID), self.mov_eval, hand, test, escala])
            fila_max_analizada += 1
            # Save the file
            wb.save(nombre_xlsx)
        else:
            wb = load_workbook(nombre_xlsx)
            ws = wb.active
            fila_max_analizada = ws.max_row
            for row in ws.iter_rows(min_row=2, max_col=5, max_row=fila_max_analizada, values_only=True):
                if row[0] == pID and row[1] == self.mov_eval and row[2] == hand and row[3] == test:
                    if row[4] == escala:
                        messagebox.showinfo(message="Esta prueba ya fue evaluada anteriormente", title="Prueba ya evaluada")
                        return
                    else:
                        choice = messagebox.askyesno(message="Esta prueba ya fue evaluada con una escala"
                                                    " diferente. ¿Desea cambiar evaluación?", title="Prueba ya evaluada")
                        if choice:
                            ws.delete_rows(filas_eliminadas, 1)
                            ws.append([str(pID), self.mov_eval, hand, test, escala])
                            fila_max_analizada += 1
                            wb.save(nombre_xlsx)
                            return
                        else:
                            return

                filas_eliminadas += 1

            ws.append([str(pID), self.mov_eval, hand, test, escala])
            fila_max_analizada += 1
            wb.save(nombre_xlsx)

    def vid_first_frame(self):
        # Show first frame of selected video
        ret, frame = self.video_cap.read()
        frame = cv2.cvtColor(resize(frame, height=613), cv2.COLOR_BGR2RGB)
        img = ImageTk.PhotoImage(image=Image.fromarray(frame))

        self.video_label.place(x=960 - frame.shape[1] // 2, y=10)
        self.video_label.config(image=img)
        self.video_label.image = img

    def vid_play(self):
        # Play selected video
        self.video_pause = False
        self.playvid_btn['state'] = tk.DISABLED
        self.vid_reproduce()

    def vid_reproduce(self):
        # Reproduce selected video
        ret, frame = self.video_cap.read()
        if ret:
            frame = cv2.cvtColor(resize(frame, height=613), cv2.COLOR_BGR2RGB)
            img = ImageTk.PhotoImage(image=Image.fromarray(frame))

            self.video_label.config(image=img)
            self.video_label.image = img
            if not self.video_pause:
                self.video_label.after(self.video_ts, self.vid_reproduce)
        else:
            # self.video_pos = 0
            self.video_cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
            self.vid_first_frame()
            self.playvid_btn['state'] = tk.NORMAL

    def vid_pause(self):
        # Pause selected video
        self.video_pause = True
        self.playvid_btn['state'] = tk.NORMAL

    def vid_first(self):
        # Go back to video first frame
        self.video_cap.set(cv2.CAP_PROP_POS_FRAMES, 0)

    def get_trajectory(self, video_path):
        # Generate csv file with hand landmarks trajectory
        if not os.path.isfile(self.csv_path):
            feature_extraction.get_trajectory(self.csv_path, video_path)
        self.process_var.set("Procesado!")


class Tab2(ttk.Frame):
    def __init__(self, container):
        super().__init__(container)
        # Create app canvas (rectangles, titles, layout)
        self.canvas = tk.Canvas(self, bg="#3A7FF6", height=720, width=1280, bd=0, highlightthickness=0, relief="ridge")
        self.canvas.place(x=0, y=0)

        # Assets
        self.btn_analyze_img = tk.PhotoImage(file=relative_to_assets("btn_analyze.png"))
        self.btn_change_img = tk.PhotoImage(file=relative_to_assets("btn_change.png"))
        self.bg_plot_img = tk.PhotoImage(file=relative_to_assets("bg_plot.png"))

        # Variables
        self.toggle1 = True
        self.toggle2 = True

        self.bandera1 = True
        self.bandera2 = True
        self.bandera3 = True
        self.bandera4 = True

        self.create_widgets()

    def create_widgets(self):
        global tab1
        # Title
        self.canvas.create_text(31.0, 44.0, anchor="nw", text="Analizar prueba", fill="#FFFFFF",
                                font=("Roboto Bold", 36 * -1))
        # Selected test: text, label, button
        self.canvas.create_text(456.0, 54.0, anchor="nw", text="Prueba seleccionada", fill="#FFFFFF",
                                font=("Roboto", 24 * -1))
        self.canvas.create_image(846.0, 67.0, image=tab1.entry_img)
        test_label = tk.Label(self, textvariable=tab1.test_var, bd=0, bg="#FFFFFF", fg="#000000", highlightthickness=0,
                              anchor="w", font=("Roboto", 13 * -1))
        test_label.place(x=705.0, y=57, width=290.0)
        analyze_btn = tk.Button(self, image=self.btn_analyze_img, borderwidth=0, highlightthickness=0,
                                command=self.analyze_video, relief="flat")
        analyze_btn.place(x=1020.0, y=43.0, width=206.0, height=47.0)
        # Graph 1: background and button
        self.canvas.create_image(337.0, 343.0, image=self.bg_plot_img)
        plot1_btn = tk.Button(self, image=self.btn_change_img, borderwidth=0, highlightthickness=0,
                              command=self.toggle_plot1, relief="flat")
        plot1_btn.place(x=395.0, y=574.0, width=206.0, height=47.0)

        # Graph 2: background and button
        self.canvas.create_image(945.0, 343.0, image=self.bg_plot_img)
        plot2_btn = tk.Button(self, image=self.btn_change_img, borderwidth=0, highlightthickness=0,
                              command=self.toggle_plot2, relief="flat")
        plot2_btn.place(x=1020.0, y=574.0, width=206.0, height=47.0)

        cursor_graph1_btn = tk.Button(self, text='Cursor 1',
                                      borderwidth=0,
                                      highlightthickness=0,
                                      command=lambda: self.cursor_graph1(),
                                      relief="flat")
        cursor_graph1_btn.place(x=80, y=574.0, width=206.0, height=47.0)

        cursor_graph2_btn = tk.Button(self, text='Cursor 2',
                                      borderwidth=0,
                                      highlightthickness=0,
                                      command=lambda: self.cursor_graph2(),
                                      relief="flat")
        cursor_graph2_btn.place(x=80, y=630.0, width=206.0, height=47.0)

        cursor_graph3_btn = tk.Button(self, text='Cursor 3',
                                      borderwidth=0,
                                      highlightthickness=0,
                                      command=lambda: self.cursor_graph3(),
                                      relief="flat")
        cursor_graph3_btn.place(x=700, y=574.0, width=206.0, height=47.0)

        cursor_graph4_btn = tk.Button(self, text='Cursor 4',
                                      borderwidth=0,
                                      highlightthickness=0,
                                      command=lambda: self.cursor_graph4(),
                                      relief="flat")
        cursor_graph4_btn.place(x=700, y=630.0, width=206.0, height=47.0)

    def analyze_video(self):
        # Get features from video
        global tab1
        csv_path = tab1.csv_path
        mov_eval = tab1.mov_eval
        fps = tab1.video_fps
        mov = tab1.movement_var.get()

        video_features = Parkinson_movements(csv_path, mov_eval, fps)
        video_features.filter_signal()
        video_features.calc_speed()

        self.t_amp, self.mov_amp1, self.amp_trend1, self.mov_amp2, self.amp_trend2 = video_features.calc_amplitude()

        self.t1, self.dedo1, self.dedo2, self.t2, self.dedo1_seg, self.dedo2_seg, self.t3, self.dedo1_vel, self.dedo2_vel = video_features.get_plot()

        self.fig1 = Figure(figsize=(5, 4), dpi=100)
        self.plot1 = self.fig1.add_subplot(111)
        self.plot1.plot(self.t1, self.dedo1)
        self.plot1.plot(self.t1, self.dedo2)
        self.plot1.set_xlabel('Tiempo (s)')
        self.plot1.set_ylabel('Amplitud')
        if mov == 'Golpeteo de dedos':
            self.plot1.set_title('Finger tapping en y - original')
            self.plot1.legend(['Pulgar', 'Índice'], loc='upper right')
        elif mov == 'Prono supinacion':
            self.plot1.set_title('Prono-supinación en x - original')
            self.plot1.legend(['Pulgar', 'Meñique'], loc='upper right')
        elif mov == 'Cierre de puño':
            self.plot1.set_title('Fist open-close en y - original')
            self.plot1.legend(['Pulgar', 'Meñique'], loc='upper right')
        self.fig1.tight_layout()

        canvasf = FigureCanvasTkAgg(self.fig1, master=self)
        canvasf.draw()
        canvasf.get_tk_widget().place(x=80.0, y=145.0)

        self.fig2 = Figure(figsize=(5, 4), dpi=100)
        self.plot2 = self.fig2.add_subplot(111)
        self.plot2.plot(self.t2, self.dedo1_seg)
        self.plot2.plot(self.t2, self.dedo2_seg)
        self.plot2.set_xlabel('Tiempo (s)')
        self.plot2.set_ylabel('Amplitud')
        if mov == 'Golpeteo de dedos':
            self.plot2.set_title('Finger tapping en y - segmentada')
            self.plot2.legend(['Pulgar', 'Índice'], loc='upper right')
        elif mov == 'Prono supinacion':
            self.plot2.set_title('Prono-supinación en x - segmentada')
            self.plot2.legend(['Pulgar', 'Meñique'], loc='upper right')
        elif mov == 'Cierre de puño':
            self.plot2.set_title('Fist open-close en y - segmentada')
            self.plot2.legend(['Pulgar', 'Meñique'], loc='upper right')
        self.fig2.tight_layout()

        # x=80, y=145

        self.fig3 = Figure(figsize=(5, 4), dpi=100)
        self.plot3 = self.fig3.add_subplot(111)
        self.plot3.plot(self.t3, self.dedo1_vel)
        self.plot3.plot(self.t3, self.dedo2_vel)
        self.plot3.set_xlabel('Tiempo (s)')
        self.plot3.set_ylabel('Velocidad')
        if mov == 'Golpeteo de dedos':
            self.plot3.set_title('Finger tapping en y - velocidad')
            self.plot3.legend(['Pulgar', 'Índice'], loc='upper right')
        elif mov == 'Prono supinacion':
            self.plot3.set_title('Prono-supinación en x - velocidad')
            self.plot3.legend(['Pulgar', 'Meñique'], loc='upper right')
        elif mov == 'Cierre de puño':
            self.plot3.set_title('Fist open-close en y - velocidad')
            self.plot3.legend(['Pulgar', 'Meñique'], loc='upper right')
        self.fig3.tight_layout()

        self.canvasf = FigureCanvasTkAgg(self.fig3, master=self)
        self.canvasf.draw()
        self.canvasf.get_tk_widget().place(x=690.0, y=145.0)
        # x=680, y=145

        self.fig4 = Figure(figsize=(5, 4), dpi=100)
        self.plot4 = self.fig4.add_subplot(111)
        self.plot4.plot(self.t_amp, self.mov_amp1, 'tab:blue')
        self.plot4.plot(self.t_amp, self.amp_trend1, color='tab:blue', linestyle=':')
        self.plot4.plot(self.t_amp, self.mov_amp2, color='tab:orange')
        self.plot4.plot(self.t_amp, self.amp_trend2, color='tab:orange', linestyle=':')
        self.plot4.set_title("Tendencias de amplitud de las señales - " + mov)
        self.plot4.set_xlabel('Tiempo(s)')
        self.plot4.set_ylabel('Amplitud')
        if mov == 'Golpeteo de dedos':
            self.plot4.legend(['Pulgar', "Tendencia pulgar", 'Índice', "Tendencia indice"], loc='upper right')
        elif mov == 'Prono supinacion':
            self.plot4.legend(['Pulgar', "Tendencia pulgar", 'Meñique', "Tendencia meñique"], loc='upper right')
        elif mov == 'Cierre de puño':
            # plot4.set_title("Falta arreglar fist open close en extraccion!!!") #No se si aun falta arreglar, toca revisar
            self.plot4.legend(['Pulgar', 'Meñique'], loc='upper right')
        self.fig4.tight_layout()

    def toggle_plot1(self):
        if self.toggle1:
            canvasf = FigureCanvasTkAgg(self.fig2, master=self)
            canvasf.draw()
            canvasf.get_tk_widget().place(x=80.0, y=145.0)
        else:
            canvasf = FigureCanvasTkAgg(self.fig1, master=self)
            canvasf.draw()
            canvasf.get_tk_widget().place(x=80.0, y=145.0)
        self.toggle1 = not self.toggle1

    def toggle_plot2(self):
        if self.toggle2:
            canvasf = FigureCanvasTkAgg(self.fig4, master=self)
            canvasf.draw()
            canvasf.get_tk_widget().place(x=690.0, y=145.0)
        else:
            canvasf = FigureCanvasTkAgg(self.fig3, master=self)
            canvasf.draw()
            canvasf.get_tk_widget().place(x=690.0, y=145.0)
        self.toggle2 = not self.toggle2

    def cursor_graph1(self):
        if self.bandera1:
            self.cursor1 = SnaptoCursor(self.plot1, self.t1, self.dedo1, "crimson")
            self.cid1 = self.fig1.canvas.mpl_connect('motion_notify_event', self.cursor1.mouse_move)
            self.cursor3 = SnaptoCursor(self.plot2, self.t2, self.dedo1_seg, "crimson")
            self.cid3 = self.fig2.canvas.mpl_connect('motion_notify_event', self.cursor3.mouse_move)
        else:
            self.cursor1.delete_cursor()
            self.cursor3.delete_cursor()
            self.fig1.canvas.mpl_disconnect(self.cid1)
            self.fig2.canvas.mpl_disconnect(self.cid3)
        self.bandera1 = not self.bandera1

    def cursor_graph2(self):
        if self.bandera2:
            self.cursor2 = SnaptoCursor(self.plot1, self.t1, self.dedo2, "blue")
            self.cid2 = self.fig1.canvas.mpl_connect('motion_notify_event', self.cursor2.mouse_move)
            self.cursor4 = SnaptoCursor(self.plot2, self.t2, self.dedo2_seg, "blue")
            self.cid4 = self.fig2.canvas.mpl_connect('motion_notify_event', self.cursor4.mouse_move)
        else:
            self.cursor2.delete_cursor()
            self.cursor4.delete_cursor()
            self.fig1.canvas.mpl_disconnect(self.cid2)
            self.fig2.canvas.mpl_disconnect(self.cid4)
        self.bandera2 = not self.bandera2

    def cursor_graph3(self):
        if self.bandera3:
            self.cursor5 = SnaptoCursor(self.plot3, self.t3, self.dedo1_vel, "crimson")
            self.cid5 = self.fig3.canvas.mpl_connect('motion_notify_event', self.cursor5.mouse_move)
            self.cursor7 = SnaptoCursor(self.plot4, self.t_amp, self.mov_amp1, "crimson")
            self.cid7 = self.fig4.canvas.mpl_connect('motion_notify_event', self.cursor7.mouse_move)
        else:
            self.cursor5.delete_cursor()
            self.cursor7.delete_cursor()
            self.fig3.canvas.mpl_disconnect(self.cid5)
            self.fig4.canvas.mpl_disconnect(self.cid7)
        self.bandera3 = not self.bandera3

    def cursor_graph4(self):
        if self.bandera4:
            self.cursor6 = SnaptoCursor(self.plot3, self.t3, self.dedo2_vel, "blue")
            self.cid6 = self.fig3.canvas.mpl_connect('motion_notify_event', self.cursor6.mouse_move)
            self.cursor8 = SnaptoCursor(self.plot4, self.t_amp, self.mov_amp2, "blue")
            self.cid8 = self.fig4.canvas.mpl_connect('motion_notify_event', self.cursor8.mouse_move)
        else:
            self.cursor6.delete_cursor()
            self.cursor8.delete_cursor()
            self.fig3.canvas.mpl_disconnect(self.cid6)
            self.fig4.canvas.mpl_disconnect(self.cid8)
        self.bandera4 = not self.bandera4


def relative_to_assets(path: str) -> Path:
    global ASSETS_PATH
    return ASSETS_PATH / Path(path)


def get_moveval(mov, hand):
    if mov == 'Golpeteo de dedos':
        mov_eval = 'fingertap'
    elif mov == 'Prono supinacion':
        mov_eval = 'pronosup'
    else:
        mov_eval = 'fist'

    if hand == 'Derecha':
        hand_eval = '_r'
    else:
        hand_eval = '_l'

    return mov_eval, hand_eval


if __name__ == '__main__':
    # Define global variables
    ASSETS_PATH = Path(__file__).parent / Path("./assets")
    CREDS_PATH = Path(__file__).parent / Path("./serviceAccountKey.json")

    # Connect to FireBase
    cred = credentials.Certificate(CREDS_PATH)
    firebase_admin.initialize_app(cred, {'storageBucket': 'parkinsondata.appspot.com'})
    bucket = storage.bucket()

    app = App()
    tab1 = Tab1(app.tabControl)
    tab2 = Tab2(app.tabControl)

    app.tabControl.add(tab1, text='Video')
    app.tabControl.add(tab2, text='Gráficas')
    app.tabControl.pack(expand=1, fill='both')
    app.mainloop()
